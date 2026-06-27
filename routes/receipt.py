from flask import Blueprint, render_template, request, send_file
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from extensions import db
from models import Ledger, AccountSubject, PaymentAccount

receipt_bp = Blueprint('receipt', __name__)


@receipt_bp.route('/receipt')
def receipt_page():
    today  = datetime.today()
    year   = int(request.args.get('year',  today.year))
    month  = int(request.args.get('month', today.month))
    rtype  = request.args.get('receipt_type', '').strip()
    pa_id  = request.args.get('payment_account_id', '').strip()
    prefix = f"{year}-{month:02d}"

    query = (Ledger.query
             .filter_by(type='출금')
             .filter(db.func.strftime('%Y-%m', Ledger.date) == prefix))
    if rtype:
        query = query.filter(Ledger.receipt_type == rtype)
    if pa_id:
        query = query.filter(Ledger.payment_account_id == pa_id)

    entries = query.order_by(Ledger.date, Ledger.id).all()

    # 증빙유형별 집계
    receipt_sums = {}
    for e in entries:
        rt = e.receipt_type or '없음'
        receipt_sums[rt] = receipt_sums.get(rt, 0) + e.amount

    # 결제수단별 집계
    account_sums = {}
    for e in entries:
        an = e.payment_account.name if e.payment_account else '미분류'
        account_sums[an] = account_sums.get(an, 0) + e.amount

    total = sum(e.amount for e in entries)
    years = list(range(today.year - 1, today.year + 2))
    accounts = PaymentAccount.query.filter_by(is_active=True).order_by(PaymentAccount.id).all()

    return render_template('receipt.html',
                           entries=entries,
                           receipt_sums=receipt_sums,
                           account_sums=account_sums,
                           total=total,
                           year=year, month=month, years=years,
                           months=range(1, 13),
                           rtype=rtype, pa_id=pa_id,
                           accounts=accounts)


# ── 엑셀 다운로드 ─────────────────────────
@receipt_bp.route('/receipt/export')
def receipt_export():
    year   = int(request.args.get('year',  datetime.today().year))
    month  = int(request.args.get('month', datetime.today().month))
    rtype  = request.args.get('receipt_type', '').strip()
    pa_id  = request.args.get('payment_account_id', '').strip()
    prefix = f"{year}-{month:02d}"

    from models import Ledger, PaymentAccount
    query = (Ledger.query
             .filter_by(type='출금')
             .filter(db.func.strftime('%Y-%m', Ledger.date) == prefix))
    if rtype:
        query = query.filter(Ledger.receipt_type == rtype)
    if pa_id:
        query = query.filter(Ledger.payment_account_id == pa_id)
    entries = query.order_by(Ledger.date, Ledger.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}년{month}월 지출증빙"

    # 스타일 정의
    hdr_font  = Font(name='맑은 고딕', bold=True, size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864', end_color='1F3864')
    hdr_align = Alignment(horizontal='center', vertical='center')
    hdr_font_color = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')

    body_font  = Font(name='맑은 고딕', size=10)
    center     = Alignment(horizontal='center', vertical='center')
    right      = Alignment(horizontal='right',  vertical='center')
    left       = Alignment(horizontal='left',   vertical='center')

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 제목 행
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"차플러스 {year}년 {month}월 지출증빙"
    title_cell.font  = Font(name='맑은 고딕', bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # 헤더 행
    headers = ['ID', '날짜', '금액', '계정과목', '증빙', '결제수단', '연결계좌/번호', '메모']
    col_widths = [12, 12, 14, 16, 12, 10, 20, 24]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = hdr_font_color
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 18

    # 데이터 행
    total = 0
    for row_idx, e in enumerate(entries, 3):
        pa = e.payment_account
        values = [
            e.id,
            str(e.date),
            e.amount,
            e.account_subject.name if e.account_subject else '',
            e.receipt_type or '없음',
            pa.type if pa else '',
            pa.number or '' if pa else '',
            e.memo or '',
        ]
        aligns = [center, center, right, left, center, center, left, left]
        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = body_font
            cell.alignment = aln
            cell.border    = border
            if col == 3:  # 금액 쉼표 포맷
                cell.number_format = '#,##0'
        # 짝수행 배경
        if row_idx % 2 == 0:
            fill = PatternFill('solid', start_color='F2F4F7', end_color='F2F4F7')
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = fill
        total += e.amount
        ws.row_dimensions[row_idx].height = 16

    # 합계 행
    total_row = len(entries) + 3
    ws.merge_cells(f'A{total_row}:B{total_row}')
    ws[f'A{total_row}'].value = '합계'
    ws[f'A{total_row}'].font = Font(name='맑은 고딕', bold=True, size=10)
    ws[f'A{total_row}'].alignment = center
    ws[f'A{total_row}'].fill = PatternFill('solid', start_color='E8ECF0', end_color='E8ECF0')
    total_cell = ws.cell(row=total_row, column=3, value=total)
    total_cell.font = Font(name='맑은 고딕', bold=True, size=10)
    total_cell.alignment = right
    total_cell.number_format = '#,##0'
    total_cell.fill = PatternFill('solid', start_color='E8ECF0', end_color='E8ECF0')
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).border = border
    ws.row_dimensions[total_row].height = 18

    ws.freeze_panes = 'A3'

    # 파일명: chaplus_yymm
    yymm = f"{str(year)[2:]}{month:02d}"
    filename = f"chaplus_{yymm}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
